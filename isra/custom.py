import frappe
from frappe import _
from frappe.utils import flt
import json
from erpnext.accounts.doctype.payment_entry.payment_entry import PaymentEntry

@frappe.whitelist()
def get_warehouse_and_location(customer):
    if not customer:
        return {"warehouse": None, "location": None, "sales_in_charge": None, "message": "Customer is required."}
    
    customer_location = frappe.db.get_value("Customer", customer, "custom_location")
    if not customer_location:
        return {"warehouse": None, "location": None, "sales_in_charge": None, "message": "Customer does not have a location specified."}

    warehouses = frappe.get_all("Warehouse", fields=["name","custom_sales_in_charge"])
    for warehouse in warehouses:
        warehouse_doc = frappe.get_doc("Warehouse", warehouse["name"])

        for row  in warehouse_doc.get("custom_location"):
            if row.location == customer_location:  # Match location in the child table
                # Return the matching warehouse and location
                return {
                    "warehouse": warehouse_doc.name,
                    "location": customer_location,
                    "sales_in_charge": warehouse_doc.custom_sales_in_charge
                }
    return None
    # 
    #     "warehouse": None,
    #     "location": None,
    #     "sales_in_charge": None,
    #     "message": f"No warehouse found for the customer's location: {customer_location}"
    # }


@frappe.whitelist()
def get_sales_invoice_item_rates(item_code, price_list=None,customer=None):
    # Fetch selling rate from Item Price
    selling_rate = frappe.db.get_value(
        'Item Price',
        filters={
            'item_code': item_code,
            'price_list': price_list,
            'selling': 1
        },
        fieldname='price_list_rate'
    )

    # Fetch previous invoice rate
    query = """
        SELECT item.rate
        FROM `tabSales Invoice Item` AS item
        JOIN `tabSales Invoice` AS invoice
        ON item.parent = invoice.name
        WHERE invoice.docstatus = 1
        AND item.item_code = %s
    """
    
    params = [item_code]
    if customer:
        query += " AND invoice.customer = %s"
        params.append(customer)
    query += " ORDER BY invoice.creation DESC LIMIT 1"
    previous_invoice_rate = frappe.db.sql(query, tuple(params), as_dict=True)

    return {
        'selling_rate': selling_rate,
        'previous_invoice_rate': previous_invoice_rate[0].rate if previous_invoice_rate else None
    }

@frappe.whitelist()
def get_sales_order_item_rates(item_code, price_list=None,customer=None):
    # Fetch selling rate from Item Price
    selling_rate = frappe.db.get_value(
        'Item Price',
        filters={
            'item_code': item_code,
            'price_list': price_list,
            'selling': 1
        },
        fieldname='price_list_rate'
    )

    # Fetch previous invoice rate
    query = """
        SELECT item.rate
        FROM `tabSales Order Item` AS item
        JOIN `tabSales Order` AS invoice
        ON item.parent = invoice.name
        WHERE invoice.docstatus = 1
        AND item.item_code = %s
    """

    params = [item_code]
    if customer:
        query += " AND invoice.customer = %s"
        params.append(customer)
    query += " ORDER BY invoice.creation DESC LIMIT 1"
    previous_invoice_rate = frappe.db.sql(query, tuple(params), as_dict=True)

    return {
        'selling_rate': selling_rate,
        'previous_invoice_rate': previous_invoice_rate[0].rate if previous_invoice_rate else None
    }


@frappe.whitelist()
def get_uoms_for_item(doctype, txt, searchfield, start, page_len, filters):
    item_code = filters.get("item_code")
    
    if not item_code:
        return []

    uoms = frappe.get_all(
        "UOM Conversion Detail",
        filters={"parent": item_code, "parenttype": "Item", "parentfield": "uoms"},
        fields=["uom"],
        order_by="idx asc"
    )

    return [[u["uom"]] for u in uoms] 

@frappe.whitelist()
def get_latest_purchase_rate(item_code):
    # ✅ First: Check latest submitted Purchase Order
    purchase_order_item = frappe.db.sql("""
        SELECT poi.rate
        FROM `tabPurchase Order Item` poi
        JOIN `tabPurchase Order` po ON po.name = poi.parent
        WHERE poi.item_code = %s AND po.docstatus = 1
        ORDER BY po.transaction_date DESC, po.creation DESC
        LIMIT 1
    """, (item_code,), as_dict=True)

    if purchase_order_item:
        return purchase_order_item[0].get("rate")

    # ❌ If not found, fallback to Purchase Invoice
    purchase_invoice_item = frappe.db.sql("""
        SELECT pii.base_rate
        FROM `tabPurchase Invoice Item` pii
        JOIN `tabPurchase Invoice` pi ON pi.name = pii.parent
        WHERE pii.item_code = %s AND pi.docstatus = 1
        ORDER BY pi.posting_date DESC, pi.creation DESC
        LIMIT 1
    """, (item_code,), as_dict=True)

    if purchase_invoice_item:
        return purchase_invoice_item[0].get("base_rate")
    return 0.0

@frappe.whitelist()
def get_stock_qty(item_code):
    return frappe.get_value("Bin", {"item_code": item_code}, "sum(actual_qty)") or 0

@frappe.whitelist()
def get_latest_purchase_price(item_code):
    purchase_invoice_item = frappe.db.sql("""
        SELECT pii.base_rate
        FROM `tabPurchase Invoice Item` pii
        JOIN `tabPurchase Invoice` pi ON pi.name = pii.parent
        WHERE pii.item_code = %s AND pi.docstatus = 1
        ORDER BY pi.posting_date DESC, pi.creation DESC
        LIMIT 1
    """, (item_code,), as_dict=True)

    if purchase_invoice_item:
        return purchase_invoice_item[0].get("base_rate")
    return 0.0

@frappe.whitelist()
def get_item_rate_and_qty(item_code):
    stock_qty = get_stock_qty(item_code)
    last_purchase_price = get_latest_purchase_price(item_code)
    return {"stock_qty": stock_qty, "last_purchase_price": last_purchase_price}

@frappe.whitelist()
def get_last_sale_qty(item_code, customer):
    # Fetch the last sales invoice item for the given item and customer
    sales_invoice_item = frappe.db.sql("""
        SELECT sii.qty, sii.parent AS invoice_name
        FROM `tabSales Invoice Item` sii
        JOIN `tabSales Invoice` si ON si.name = sii.parent
        WHERE sii.item_code = %s AND si.customer = %s AND si.docstatus = 1 AND si.return_against IS NULL
        ORDER BY si.posting_date DESC, si.creation DESC
        LIMIT 1
    """, (item_code, customer), as_dict=True)

    if not sales_invoice_item:
        return 0.0

    last_sale_qty = sales_invoice_item[0].get("qty")
    invoice_name = sales_invoice_item[0].get("invoice_name")

    # Check if there is a return against the last sales invoice
    return_entry = frappe.db.sql("""
        SELECT sii.qty
        FROM `tabSales Invoice Item` sii
        JOIN `tabSales Invoice` si ON si.name = sii.parent
        WHERE si.return_against = %s AND sii.item_code = %s AND si.docstatus = 1
    """, (invoice_name, item_code), as_dict=True)
    if return_entry:
        returned_qty = sum(entry.get("qty") for entry in return_entry)
        last_sale_qty += returned_qty

    return last_sale_qty

@frappe.whitelist()
def download_denomination_report(data, account_amount:float):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from io import BytesIO 
    if isinstance(data, str):
        data = json.loads(data)
    denominations = []
    total = 0
    for i in data:
        row_total = float(i['denomination']) * float(i['pieces'])
        total += row_total
        denominations.append((f"{i['denomination']} X {i['pieces']}", row_total))
    cash_short = total - account_amount  # -100

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Report"

    # Headers
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 10
    ws["A1"] = "Denomination"
    ws["B1"] = "Amount"
    ws["A1"].font = ws["B1"].font = Font(bold=True)
    ws["A1"].alignment = ws["B1"].alignment = Alignment(horizontal="center")

    # Data rows
    row = 2
    for denom, amt in denominations:
        ws[f"A{row}"] = denom
        ws[f"B{row}"] = amt
        row += 1

    row += 1
    # Totals
    ws[f"A{row}"] = "Total Cash In Hand"
    ws[f"B{row}"] = total
    ws[f"A{row}"].font = ws[f"B{row}"].font = Font(bold=True)

    row += 1
    ws[f"A{row}"] = "Total Amount In Account"
    ws[f"B{row}"] = account_amount
    ws[f"A{row}"].font = ws[f"B{row}"].font = Font(bold=True)

    row += 1
    ws[f"A{row}"] = "Cash Short"
    ws[f"B{row}"] = cash_short
    ws[f"A{row}"].font = ws[f"B{row}"].font = Font(bold=True)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Prepare Frappe response
    frappe.local.response.filename = "denomination_report.xlsx"
    frappe.local.response.filecontent = output.getvalue()
    frappe.local.response.type = "download"